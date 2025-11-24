import base64
import logging
import secrets
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
SCREENSHOT_DIR = BASE_DIR / "screenshots"
TICKETS_FILE = BASE_DIR / "tickets.xlsx"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("citi_ghost_feedback")

app = Flask(__name__)


def ensure_storage():
    """Ensure screenshots directory and workbook exist before serving requests."""
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    if not TICKETS_FILE.exists():
        logger.info("tickets.xlsx not found. Creating a new workbook.")
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Tickets"
        sheet.append(
            [
                "Ticket ID",
                "Role",
                "URL",
                "Description",
                "Screenshot Path",
                "User Agent",
                "Timestamp",
                "Issue Type",
                "Priority",
                "Category",
                "Archive Status",
            ]
        )
        wb.save(TICKETS_FILE)


ensure_storage()


def load_sheet():
    wb = load_workbook(TICKETS_FILE)
    return wb, wb.active


def save_screenshot(data_url: str) -> str:
    """Save a base64 screenshot to disk and return the relative path."""
    try:
        header, encoded = data_url.split(",", 1)
    except ValueError:
        raise ValueError("Invalid screenshot payload")

    filename = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(4)}.png"
    path = SCREENSHOT_DIR / filename
    image_bytes = base64.b64decode(encoded)
    with open(path, "wb") as f:
        f.write(image_bytes)
    logger.info("Saved screenshot to %s", path)
    return str(path.relative_to(BASE_DIR))


def append_ticket(role, url, description, screenshot_path, user_agent, issue_type="", priority="", category=""):
    wb, sheet = load_sheet()
    ticket_id = sheet.max_row  # header is row 1, so next row index == ticket id
    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    
    # Add metadata to description if available
    full_description = description
    if issue_type or priority or category:
        metadata = []
        if issue_type:
            metadata.append(f"[{issue_type}]")
        if priority:
            metadata.append(f"Priority: {priority}")
        if category:
            metadata.append(f"Tag: {category}")
        full_description = f"{' '.join(metadata)}\n\n{description}"
    
    # Ensure header has all columns
    headers = [cell.value for cell in sheet[1]]
    if len(headers) < 11:
        # Add missing header columns
        while len(headers) < 11:
            col_idx = len(headers) + 1
            if col_idx == 8:
                sheet.cell(row=1, column=col_idx, value="Issue Type")
            elif col_idx == 9:
                sheet.cell(row=1, column=col_idx, value="Priority")
            elif col_idx == 10:
                sheet.cell(row=1, column=col_idx, value="Category")
            elif col_idx == 11:
                sheet.cell(row=1, column=col_idx, value="Archive Status")
            headers.append(sheet.cell(row=1, column=col_idx).value)
    
    sheet.append(
        [
            ticket_id,
            role,
            url,
            full_description,
            screenshot_path or "",
            user_agent,
            timestamp,
            issue_type or "",
            priority or "",
            category or "",
            "active",  # Default archive status
        ]
    )
    wb.save(TICKETS_FILE)
    logger.info("Stored ticket %s", ticket_id)
    return ticket_id, timestamp


@app.after_request
def apply_cors(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return response


@app.route("/screenshots/<path:filename>")
def serve_screenshot(filename):
    """Serve screenshot files."""
    return send_from_directory(SCREENSHOT_DIR, filename)


@app.route("/get-feedback", methods=["GET", "OPTIONS"])
def get_feedback():
    """Retrieve all feedback tickets from the Excel file."""
    if request.method == "OPTIONS":
        return ("", 204)

    try:
        wb, sheet = load_sheet()
        tickets = []

        # Skip header row (row 1)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If ticket ID exists
                # Handle old format (without archive status) - default to "active"
                archive_status = row[10] if len(row) > 10 else "active"
                tickets.append({
                    "ticket_id": row[0],
                    "role": row[1] or "Client",
                    "url": row[2] or "",
                    "description": row[3] or "",
                    "screenshot_path": row[4] or "",
                    "user_agent": row[5] or "",
                    "timestamp": row[6] or "",
                    "issue_type": row[7] if len(row) > 7 else "",
                    "priority": row[8] if len(row) > 8 else "",
                    "category": row[9] if len(row) > 9 else "",
                    "archive_status": archive_status,
                })

        # Reverse to show newest first
        tickets.reverse()
        return jsonify({"status": "success", "tickets": tickets, "count": len(tickets)})
    except Exception as exc:
        logger.exception("Unable to load tickets: %s", exc)
        return jsonify({"status": "error", "message": "Failed to load feedback"}), 500


@app.route("/submit-feedback", methods=["POST", "OPTIONS"])
def submit_feedback():
    if request.method == "OPTIONS":
        return ("", 204)

    payload = request.get_json(force=True, silent=True) or {}
    role = payload.get("role", "").strip() or "Client"
    description = payload.get("description", "").strip()
    url = payload.get("url", "").strip()
    user_agent = payload.get("userAgent", "Unknown")
    screenshot_data = payload.get("screenshot")
    issue_type = payload.get("issueType", "").strip()
    priority = payload.get("priority", "").strip()
    category = payload.get("category", "").strip()

    if not description:
        logger.warning("Description missing from payload")
        return jsonify({"status": "error", "message": "Description is required"}), 400

    screenshot_path = ""
    if screenshot_data:
        try:
            screenshot_path = save_screenshot(screenshot_data)
        except Exception as exc:  # broad to ensure request still stored
            logger.error("Failed to save screenshot: %s", exc)
            return (
                jsonify({"status": "error", "message": "Invalid screenshot data"}),
                400,
            )

    try:
        ticket_id, timestamp = append_ticket(
            role=role,
            url=url or request.headers.get("Referer", ""),
            description=description,
            screenshot_path=screenshot_path,
            user_agent=user_agent,
            issue_type=issue_type,
            priority=priority,
            category=category,
        )
    except Exception as exc:
        logger.exception("Unable to append ticket: %s", exc)
        return jsonify({"status": "error", "message": "Failed to store feedback"}), 500

    return jsonify({"status": "success", "ticket_id": ticket_id, "timestamp": timestamp})


@app.route("/archive-feedback", methods=["PUT", "OPTIONS"])
def archive_feedback():
    """Archive or unarchive a feedback ticket."""
    if request.method == "OPTIONS":
        return ("", 204)

    payload = request.get_json(force=True, silent=True) or {}
    ticket_id = payload.get("ticket_id")
    archive_status = payload.get("status", "archived")  # archived, resolved, closed, active

    if not ticket_id:
        return jsonify({"status": "error", "message": "Ticket ID is required"}), 400

    try:
        wb, sheet = load_sheet()
        
        # Find the ticket by ID
        found = False
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == ticket_id:
                # Ensure Archive Status column exists
                if sheet.max_column < 11:
                    sheet.cell(row=1, column=11, value="Archive Status")
                sheet.cell(row=row_idx, column=11, value=archive_status)
                found = True
                break

        if not found:
            return jsonify({"status": "error", "message": "Ticket not found"}), 404

        wb.save(TICKETS_FILE)
        logger.info("Updated ticket %s archive status to %s", ticket_id, archive_status)
        return jsonify({"status": "success", "ticket_id": ticket_id, "archive_status": archive_status})
    except Exception as exc:
        logger.exception("Unable to update ticket archive status: %s", exc)
        return jsonify({"status": "error", "message": "Failed to update archive status"}), 500


if __name__ == "__main__":
    ensure_storage()
    app.run(host="127.0.0.1", port=5000, debug=True)
